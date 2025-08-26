"""
企业政策预审系统测试工具
专注于测试企业数据和企业类型政策的预审
使用preaudit-single接口，生成详细的预审结果Excel报告
"""

import requests
import json
import time
import os
from typing import List, Dict
from datetime import datetime
import numpy as np
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class EnterprisePreAuditTester:
    """企业政策预审测试类 - 专注于企业政策预审"""
    
    def __init__(self, base_url: str = "http://127.0.0.1:8084"):
        self.base_url = base_url
        self.audit_matrix = []  # 存储所有预审结果
        
        # 验证API服务
        self.verify_api_service()
    
    def verify_api_service(self):
        """验证API服务是否可用"""
        try:
            response = requests.get(f"{self.base_url}/health", timeout=5)
            if response.status_code == 200:
                data = response.json()
                service_info = data.get('message', '')
                logger.info(f"✅ 成功连接到预审API服务: {service_info}")
            else:
                logger.warning(f"⚠️  预审API服务响应异常: {response.status_code}")
        except Exception as e:
            logger.error(f"❌ 无法连接到预审API服务 {self.base_url}: {e}")
            logger.error("请确保预审API服务正在运行！")
    
    def load_test_data(self, policy_folder: str = "policy_new", enterprise_folder: str = "enterprise_dataset", 
                       policy_limit: int = 50, enterprise_limit: int = 50):
        """加载测试数据 - 只加载企业政策"""
        logger.info(f"📂 开始加载测试数据")
        logger.info(f"   政策文件夹: {policy_folder} (限制: {policy_limit})")
        logger.info(f"   企业文件夹: {enterprise_folder} (限制: {enterprise_limit})")
        
        # 加载所有政策数据
        all_policies = []
        if os.path.exists(policy_folder):
            policy_files = sorted([f for f in os.listdir(policy_folder) if f.endswith('.json')])
            logger.info(f"发现 {len(policy_files)} 个政策文件")
            
            for file in policy_files[:policy_limit]:
                try:
                    with open(os.path.join(policy_folder, file), 'r', encoding='utf-8') as f:
                        policy_data = json.load(f)
                        all_policies.append(policy_data)
                        logger.debug(f"✅ 加载政策: {file}")
                except Exception as e:
                    logger.warning(f"❌ 加载政策文件失败: {file} - {e}")
        else:
            logger.error(f"❌ 政策文件夹不存在: {policy_folder}")
            return [], []
        
        # 显示所有政策类型统计
        logger.info(f"\n📊 政策类型分析 (共 {len(all_policies)} 个政策):")
        policy_type_stats = {}
        for policy in all_policies:
            policy_id = policy.get('政策编号', 'Unknown')
            policy_type = policy.get('类型', 'Unknown')
            title = policy.get('标题', '无标题')[:30]  # 只显示前30个字符
            
            if policy_type not in policy_type_stats:
                policy_type_stats[policy_type] = []
            policy_type_stats[policy_type].append(policy_id)
            
            logger.info(f"   {policy_id}: 类型='{policy_type}' | {title}")
        
        # 显示类型统计
        logger.info(f"\n📈 政策类型统计:")
        for policy_type, policies_list in policy_type_stats.items():
            logger.info(f"   '{policy_type}': {len(policies_list)} 个")
        
        # 筛选企业政策
        enterprise_policies = []
        for policy in all_policies:
            if self.is_enterprise_policy(policy):
                enterprise_policies.append(policy)
        
        if not enterprise_policies:
            logger.error(f"❌ 没有找到任何企业政策！")
            logger.error("请检查政策文件中的'类型'字段是否正确设置为'企业'")
            return [], []
        
        # 加载企业数据
        enterprises = []
        if os.path.exists(enterprise_folder):
            enterprise_files = sorted([f for f in os.listdir(enterprise_folder) if f.endswith('.json')])
            logger.info(f"\n发现 {len(enterprise_files)} 个企业文件")
            
            for file in enterprise_files[:enterprise_limit]:
                try:
                    with open(os.path.join(enterprise_folder, file), 'r', encoding='utf-8') as f:
                        enterprise_data = json.load(f)
                        enterprises.append(enterprise_data)
                        logger.debug(f"✅ 加载企业: {file}")
                except Exception as e:
                    logger.warning(f"❌ 加载企业文件失败: {file} - {e}")
        else:
            logger.error(f"❌ 企业文件夹不存在: {enterprise_folder}")
            return [], []
        
        if not enterprises:
            logger.error("❌ 没有成功加载任何企业数据")
            return [], []
        
        logger.info(f"\n✅ 数据加载完成:")
        logger.info(f"   企业政策: {len(enterprise_policies)} 个")
        logger.info(f"   企业数据: {len(enterprises)} 个")
        logger.info(f"   总预审测试: {len(enterprise_policies) * len(enterprises)} 次")
        
        # 显示筛选出的企业政策
        logger.info(f"\n📋 筛选出的企业政策:")
        for policy in enterprise_policies:
            policy_id = policy.get('政策编号', 'Unknown')
            title = policy.get('标题', '无标题')[:40]
            logger.info(f"   {policy_id}: {title}")
        
        return enterprise_policies, enterprises
    
    def is_enterprise_policy(self, policy_data: Dict) -> bool:
        """严格判断是否为企业政策"""
        try:
            policy_id = policy_data.get("政策编号", "Unknown")
            
            # 1. 检查主要类型字段
            policy_type = policy_data.get("类型", "").strip()
            if policy_type == "企业":
                logger.debug(f"✅ 企业政策 (类型字段): {policy_id}")
                return True
            
            # 2. 检查备用类型字段
            policy_type_alt = policy_data.get("政策类型", "").strip() 
            if policy_type_alt == "企业":
                logger.debug(f"✅ 企业政策 (政策类型字段): {policy_id}")
                return True
            
            # 3. 如果明确标记为其他类型，直接排除
            non_enterprise_types = ["个人", "人才", "学生", "居民", "市民"]
            if policy_type in non_enterprise_types or policy_type_alt in non_enterprise_types:
                logger.debug(f"❌ 非企业政策 (明确类型): {policy_id} - {policy_type}")
                return False
            
            # 4. 检查标题关键词
            title = policy_data.get("标题", "")
            
            # 企业关键词
            enterprise_keywords = ["企业", "公司", "法人", "商户", "经营", "创业", "营业执照", "注册资本", "工商", "商业"]
            # 排除关键词
            exclude_keywords = ["个人", "人才", "学生", "居民", "就业", "求职", "毕业", "户籍", "市民"]
            
            # 如果标题包含排除关键词，排除
            if any(keyword in title for keyword in exclude_keywords):
                logger.debug(f"❌ 非企业政策 (标题排除): {policy_id}")
                return False
            
            # 如果标题包含企业关键词，包含
            if any(keyword in title for keyword in enterprise_keywords):
                logger.debug(f"✅ 企业政策 (标题匹配): {policy_id}")
                return True
            
            # 5. 检查条件字段
            enterprise_field_count = 0
            personal_field_count = 0
            
            conditions = policy_data.get("条件", {})
            rules = conditions.get("规则", [])
            
            # 企业特有字段
            enterprise_fields = ["注册资本", "经营时间", "企业规模", "注册地", "法人年龄", 
                               "营业执照", "年营业额", "员工人数", "缴纳社保", "法定代表人"]
            # 个人特有字段
            personal_fields = ["年龄", "学历", "户籍", "婚姻状况", "工作年限", "职业", "收入", "身份证"]
            
            def count_fields(rules_list):
                nonlocal enterprise_field_count, personal_field_count
                for rule in rules_list:
                    if isinstance(rule, dict):
                        field = rule.get("字段", "")
                        
                        # 递归检查嵌套规则
                        if "规则" in rule and isinstance(rule["规则"], list):
                            count_fields(rule["规则"])
                        
                        if field in enterprise_fields:
                            enterprise_field_count += 1
                        elif field in personal_fields:
                            personal_field_count += 1
            
            count_fields(rules)
            
            # 如果有个人字段但没有企业字段，排除
            if personal_field_count > 0 and enterprise_field_count == 0:
                logger.debug(f"❌ 非企业政策 (个人字段): {policy_id}")
                return False
            
            # 如果有企业字段，包含
            if enterprise_field_count > 0:
                logger.debug(f"✅ 企业政策 (企业字段): {policy_id}")
                return True
            
            # 默认情况：如果类型不明确，不包含
            logger.debug(f"❌ 类型不明确，排除: {policy_id}")
            return False
                
        except Exception as e:
            logger.warning(f"判断政策类型出错 {policy_id}: {e}")
            return False
    
    def test_single_preaudit(self, enterprise_data: Dict, policy_data: Dict):
        """测试单个企业和政策的预审"""
        start_time = time.time()
        
        try:
            enterprise_id = enterprise_data.get('企业ID', 'Unknown')
            policy_id = policy_data.get('政策编号', 'Unknown')
            policy_type = policy_data.get('类型', 'Unknown')
            
            # 转换为预审API所需的格式
            payload = {
                "enterprise": enterprise_data,  # 企业数据作为enterprise传递
                "policy": policy_data
            }
            
            response = requests.post(
                f"{self.base_url}/preaudit-single",
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=30
            )
            end_time = time.time()
            
            if response.status_code == 200:
                result_data = response.json()
                audit_result = result_data.get('result', 0)  # 0或1
                
                return {
                    'enterprise_id': enterprise_id,
                    'policy_id': policy_id,
                    'policy_type': policy_type,
                    'audit_result': audit_result,
                    'response_time': round((end_time - start_time) * 1000, 2),
                    'success': True
                }
            else:
                logger.warning(f"API调用失败: {enterprise_id} x {policy_id} - {response.status_code}")
                return {
                    'enterprise_id': enterprise_id,
                    'policy_id': policy_id,
                    'policy_type': policy_type,
                    'audit_result': 0,
                    'response_time': round((end_time - start_time) * 1000, 2),
                    'success': False,
                    'error': f"HTTP {response.status_code}"
                }
                
        except Exception as e:
            return {
                'enterprise_id': enterprise_data.get('企业ID', 'Unknown'),
                'policy_id': policy_data.get('政策编号', 'Unknown'),
                'policy_type': policy_data.get('类型', 'Unknown'),
                'audit_result': 0,
                'response_time': 0,
                'success': False,
                'error': str(e)
            }
    
    def generate_audit_matrix(self, enterprises: List[Dict], policies: List[Dict]):
        """生成企业-政策预审结果矩阵"""
        logger.info(f"\n🔍 开始生成预审结果矩阵")
        logger.info(f"   {len(enterprises)} 个企业 × {len(policies)} 个企业政策")
        
        audit_matrix = []
        total_tests = len(enterprises) * len(policies)
        current_test = 0
        start_time = time.time()
        
        for enterprise in enterprises:
            enterprise_id = enterprise.get('企业ID')
            enterprise_audits = {
                'enterprise_id': enterprise_id,
                'enterprise_info': enterprise,
                'policy_audits': []
            }
            
            logger.info(f"🏢 测试企业 {enterprise_id} 的政策预审...")
            
            for policy in policies:
                current_test += 1
                policy_id = policy.get('政策编号')
                
                # 调用预审测试
                result = self.test_single_preaudit(enterprise, policy)
                audit_result = result.get('audit_result', 0)
                success = result.get('success', False)
                
                enterprise_audits['policy_audits'].append({
                    'policy_id': policy_id,
                    'policy_info': policy,
                    'audit_result': audit_result,  # 0或1
                    'success': success,
                    'response_time': result.get('response_time', 0)
                })
                
                # 显示进度和结果
                status = "✅通过" if audit_result == 1 else "❌未通过"
                if not success:
                    status = "⚠️失败"
                logger.info(f"   {policy_id}: {status} ({audit_result}) ({current_test}/{total_tests})")
                
                # 避免请求过于频繁
                time.sleep(0.05)
            
            audit_matrix.append(enterprise_audits)
        
        end_time = time.time()
        
        # 统计结果
        successful_tests = sum(1 for em in audit_matrix for pa in em['policy_audits'] if pa['success'])
        failed_tests = total_tests - successful_tests
        passed_audits = sum(1 for em in audit_matrix for pa in em['policy_audits'] if pa['audit_result'] == 1)
        avg_response_time = np.mean([pa['response_time'] for em in audit_matrix for pa in em['policy_audits'] if pa['response_time'] > 0])
        
        logger.info(f"\n✅ 预审结果矩阵生成完成!")
        logger.info(f"   总测试数: {total_tests}")
        logger.info(f"   API成功: {successful_tests} | API失败: {failed_tests}")
        logger.info(f"   预审通过: {passed_audits} | 预审未通过: {total_tests - passed_audits}")
        logger.info(f"   API成功率: {(successful_tests/total_tests*100):.1f}%")
        logger.info(f"   预审通过率: {(passed_audits/total_tests*100):.1f}%")
        logger.info(f"   平均响应时间: {avg_response_time:.1f}ms")
        logger.info(f"   总耗时: {(end_time-start_time):.1f}秒")
        
        self.audit_matrix = audit_matrix
        return audit_matrix
    
    def format_policy_conditions(self, policy: Dict) -> str:
        """格式化政策条件为可读字符串"""
        def parse_rules(rules, parent_logic="AND"):
            if not isinstance(rules, list):
                return []
            
            conditions = []
            for rule in rules:
                if not isinstance(rule, dict):
                    continue
                
                # 处理嵌套规则
                if '规则' in rule and isinstance(rule['规则'], list):
                    nested_logic = rule.get('逻辑', 'AND')
                    nested_conditions = parse_rules(rule['规则'], nested_logic)
                    if nested_conditions:
                        connector = " 或 " if nested_logic == 'OR' else " 且 "
                        conditions.append(f"({connector.join(nested_conditions)})")
                else:
                    # 处理单个条件
                    field = rule.get('字段', '')
                    value = rule.get('值', '')
                    operator = rule.get('操作符', '')
                    desc = rule.get('描述', '')
                    
                    if desc:  # 优先使用描述
                        conditions.append(desc)
                    elif field and value is not None:
                        if operator == '<=':
                            conditions.append(f"{field}≤{value}")
                        elif operator == '>=':
                            conditions.append(f"{field}≥{value}")
                        elif operator == '=':
                            conditions.append(f"{field}={value}")
                        elif operator == '!=':
                            conditions.append(f"{field}≠{value}")
                        elif operator == 'between' and isinstance(value, list) and len(value) == 2:
                            conditions.append(f"{field}∈[{value[0]},{value[1]}]")
                        elif operator == 'contains':
                            conditions.append(f"{field}包含'{value}'")
                        elif operator == 'in' and isinstance(value, list):
                            conditions.append(f"{field}∈{value}")
                        else:
                            conditions.append(f"{field}:{value}")
            
            return conditions
        
        # 获取条件规则
        conditions = policy.get('条件', {})
        rules = conditions.get('规则', [])
        main_logic = conditions.get('逻辑', 'AND')
        
        if not rules:
            return "无条件限制"
        
        all_conditions = parse_rules(rules, main_logic)
        if not all_conditions:
            return "条件解析失败"
        
        connector = " 或 " if main_logic == 'OR' else " 且 "
        return connector.join(all_conditions)
    
    def format_enterprise_info(self, enterprise: Dict) -> str:
        """格式化企业信息为可读字符串"""
        info_parts = []
        
        # 关键字段及其单位
        key_fields = [
            ('企业ID', ''),
            ('行业', ''),
            ('注册资本', '万元'),
            ('经营时间', '年'),
            ('员工人数', '人'),
            ('年营业额', '万元'),
            ('法人年龄', '岁'),
            ('企业规模', ''),
            ('注册地', ''),
            ('纳税情况', ''),
            ('贷款情况', '')
        ]
        
        for field, unit in key_fields:
            if field in enterprise and enterprise[field] is not None:
                value = enterprise[field]
                if unit:
                    info_parts.append(f"{field}:{value}{unit}")
                else:
                    info_parts.append(f"{field}:{value}")
        
        return " | ".join(info_parts) if info_parts else "企业信息不完整"
    
    def export_to_excel(self, filename: str = None):
        """导出预审结果矩阵到Excel"""
        if not self.audit_matrix:
            logger.error("❌ 没有预审数据可导出")
            return None
        
        if filename is None:
            filename = f'企业政策预审结果矩阵_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        logger.info(f"📊 开始导出Excel报告: {filename}")
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "企业政策预审结果矩阵"
        
        policies = self.audit_matrix[0]['policy_audits']
        enterprises = self.audit_matrix
        
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        policy_font = Font(bold=True, color="FFFFFF")
        policy_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        enterprise_font = Font(bold=True)
        enterprise_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 第1行：政策编号
        ws.merge_cells('A1:B1')
        ws['A1'] = "企业信息 / 政策编号"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].border = border
        
        col = 3
        for policy_audit in policies:
            policy_id = policy_audit['policy_id']
            ws.cell(row=1, column=col, value=policy_id)
            cell = ws.cell(row=1, column=col)
            cell.font = policy_font
            cell.fill = policy_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            col += 1
        
        # 第2行：政策标题
        ws.merge_cells('A2:B2')
        ws['A2'] = "政策标题"
        ws['A2'].font = header_font
        ws['A2'].fill = header_fill
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].border = border
        
        col = 3
        for policy_audit in policies:
            policy_title = policy_audit['policy_info'].get('标题', '无标题')
            cell = ws.cell(row=2, column=col, value=policy_title)
            cell.font = Font(size=10, bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            col += 1
        
        # 第3行：政策条件
        ws.merge_cells('A3:B3')
        ws['A3'] = "政策条件"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A3'].border = border
        
        col = 3
        for policy_audit in policies:
            conditions = self.format_policy_conditions(policy_audit['policy_info'])
            cell = ws.cell(row=3, column=col, value=conditions)
            cell.font = Font(size=9)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border
            col += 1
        
        # 第4行开始：企业数据和预审结果
        row = 4
        for enterprise_audit in enterprises:
            enterprise_id = enterprise_audit['enterprise_id']
            enterprise_desc = self.format_enterprise_info(enterprise_audit['enterprise_info'])
            
            # A列：企业ID
            cell_a = ws.cell(row=row, column=1, value=enterprise_id)
            cell_a.font = enterprise_font
            cell_a.fill = enterprise_fill
            cell_a.alignment = Alignment(horizontal='center', vertical='center')
            cell_a.border = border
            
            # B列：企业信息
            cell_b = ws.cell(row=row, column=2, value=enterprise_desc)
            cell_b.font = Font(size=9)
            cell_b.fill = enterprise_fill
            cell_b.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell_b.border = border
            
            # C列开始：预审结果
            col = 3
            for policy_audit in enterprise_audit['policy_audits']:
                audit_result = policy_audit['audit_result']
                
                cell = ws.cell(row=row, column=col, value=audit_result)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                cell.font = Font(bold=True)
                
                # 根据预审结果设置颜色
                if audit_result == 1:
                    # 淡绿色 - 通过
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    cell.font = Font(color="2E7D32", bold=True)
                else:
                    # 淡红色 - 未通过
                    cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    cell.font = Font(color="C62828", bold=True)
                
                col += 1
            
            row += 1
        
        # 调整列宽和行高
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 80
        for i in range(len(policies)):
            col_letter = chr(67 + i)  # C, D, E, ...
            ws.column_dimensions[col_letter].width = 15
        
        ws.row_dimensions[2].height = 60  # 政策标题行
        ws.row_dimensions[3].height = 80  # 政策条件行
        for row_num in range(4, 4 + len(enterprises)):
            ws.row_dimensions[row_num].height = 50
        
        # 添加统计工作表
        self._add_statistics_sheet(wb, enterprises, policies)
        
        # 保存文件
        try:
            wb.save(filename)
            logger.info(f"✅ Excel报告已保存: {filename}")
            return filename
        except Exception as e:
            logger.error(f"❌ 保存Excel文件失败: {e}")
            return None
    
    def _add_statistics_sheet(self, workbook, enterprises, policies):
        """添加统计分析工作表"""
        ws = workbook.create_sheet("统计分析")
        
        # 标题
        ws['A1'] = "企业政策预审结果统计分析"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # 企业统计
        ws[f'A{row}'] = "📊 企业预审结果统计"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        headers = ["企业ID", "通过政策数", "总政策数", "通过率(%)", "响应成功率"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        for enterprise_audit in self.audit_matrix:
            enterprise_id = enterprise_audit['enterprise_id']
            results = [pa['audit_result'] for pa in enterprise_audit['policy_audits']]
            successes = [pa['success'] for pa in enterprise_audit['policy_audits']]
            
            passed_count = sum(results)
            total_count = len(results)
            pass_rate = round(passed_count / total_count * 100, 1) if total_count > 0 else 0
            success_rate = f"{sum(successes)}/{len(successes)}"
            
            ws.cell(row=row, column=1, value=enterprise_id)
            ws.cell(row=row, column=2, value=passed_count)
            ws.cell(row=row, column=3, value=total_count)
            ws.cell(row=row, column=4, value=pass_rate)
            ws.cell(row=row, column=5, value=success_rate)
            row += 1
        
        row += 2
        
        # 政策统计
        ws[f'A{row}'] = "📋 政策预审结果统计"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        headers = ["政策编号", "政策标题", "通过企业数", "通过率(%)", "响应成功率"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        row += 1
        
        # 按政策统计
        policy_stats = {}
        for enterprise_audit in self.audit_matrix:
            for policy_audit in enterprise_audit['policy_audits']:
                policy_id = policy_audit['policy_id']
                if policy_id not in policy_stats:
                    policy_stats[policy_id] = {
                        'title': policy_audit['policy_info'].get('标题', ''),
                        'results': [],
                        'successes': []
                    }
                policy_stats[policy_id]['results'].append(policy_audit['audit_result'])
                policy_stats[policy_id]['successes'].append(policy_audit['success'])
        
        for policy_id, stats in policy_stats.items():
            results = stats['results']
            successes = stats['successes']
            
            passed_count = sum(results)
            total_count = len(results)
            pass_rate = round(passed_count / total_count * 100, 1) if total_count > 0 else 0
            success_rate = f"{sum(successes)}/{len(successes)}"
            
            ws.cell(row=row, column=1, value=policy_id)
            ws.cell(row=row, column=2, value=stats['title'][:50])  # 限制长度
            ws.cell(row=row, column=3, value=passed_count)
            ws.cell(row=row, column=4, value=pass_rate)
            ws.cell(row=row, column=5, value=success_rate)
            row += 1
        
        # 总体统计
        row += 2
        ws[f'A{row}'] = "📈 总体统计"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        all_results = [pa['audit_result'] for em in self.audit_matrix for pa in em['policy_audits']]
        all_successes = [pa['success'] for em in self.audit_matrix for pa in em['policy_audits']]
        
        stats_data = [
            ("总预审测试", len(all_results)),
            ("API成功次数", sum(all_successes)),
            ("API成功率", f"{sum(all_successes)/len(all_successes)*100:.1f}%"),
            ("预审通过次数", sum(all_results)),
            ("预审通过率", f"{sum(all_results)/len(all_results)*100:.1f}%"),
            ("预审未通过次数", len(all_results) - sum(all_results)),
            ("平均响应时间", f"{np.mean([pa['response_time'] for em in self.audit_matrix for pa in em['policy_audits'] if pa['response_time'] > 0]):.1f}ms"),
        ]
        
        for stat_name, stat_value in stats_data:
            ws.cell(row=row, column=1, value=stat_name).font = Font(bold=True)
            ws.cell(row=row, column=2, value=stat_value)
            row += 1
        
        # 调整列宽
        for i, width in enumerate([15, 40, 15, 15, 15], 1):
            ws.column_dimensions[chr(64 + i)].width = width
    
    def run_enterprise_preaudit_test(self, policy_limit: int = 8, enterprise_limit: int = 50):
        """运行企业政策预审测试"""
        logger.info("🚀 开始企业政策预审系统测试")
        logger.info("="*60)
        test_start_time = time.time()
        
        # 1. 加载测试数据
        policies, enterprises = self.load_test_data(
            policy_limit=policy_limit, 
            enterprise_limit=enterprise_limit
        )
        
        if not policies or not enterprises:
            logger.error("❌ 测试数据加载失败，无法继续测试")
            return None
        
        # 2. 生成预审结果矩阵
        logger.info("\n" + "="*60)
        logger.info("🔍 开始生成企业政策预审结果矩阵")
        logger.info("="*60)
        
        audit_matrix = self.generate_audit_matrix(enterprises, policies)
        
        if not audit_matrix:
            logger.error("❌ 预审结果矩阵生成失败")
            return None
        
        # 3. 导出Excel报告
        logger.info("\n" + "="*60)
        logger.info("📊 导出Excel报告")
        logger.info("="*60)
        
        excel_filename = self.export_to_excel()
        
        test_end_time = time.time()
        
        # 4. 生成测试总结
        logger.info("\n" + "="*60)
        logger.info("📋 测试总结")
        logger.info("="*60)
        
        all_results = [pa['audit_result'] for em in audit_matrix for pa in em['policy_audits']]
        all_successes = [pa['success'] for em in audit_matrix for pa in em['policy_audits']]
        total_time = test_end_time - test_start_time
        
        logger.info(f"✅ 测试完成！")
        logger.info(f"   📊 数据规模: {len(policies)} 个企业政策 × {len(enterprises)} 个企业")
        logger.info(f"   🔢 总测试数: {len(all_results)} 次")
        logger.info(f"   ✅ API成功率: {sum(all_successes)}/{len(all_successes)} ({sum(all_successes)/len(all_successes)*100:.1f}%)")
        logger.info(f"   📈 预审通过率: {sum(all_results)}/{len(all_results)} ({sum(all_results)/len(all_results)*100:.1f}%)")
        logger.info(f"   ⏱️  总耗时: {total_time:.1f} 秒")
        if excel_filename:
            logger.info(f"   📄 Excel报告: {excel_filename}")
        
        logger.info("="*60)
        logger.info("🎉 企业政策预审测试完成！")
        logger.info("   请查看Excel文件获取详细的预审结果矩阵和统计分析")
        logger.info("   绿色表示预审通过(1)，红色表示预审未通过(0)")
        logger.info("="*60)
        
        return {
            'policies_count': len(policies),
            'enterprises_count': len(enterprises),
            'total_tests': len(all_results),
            'api_success_rate': sum(all_successes)/len(all_successes)*100,
            'audit_pass_rate': sum(all_results)/len(all_results)*100,
            'total_time': total_time,
            'excel_file': excel_filename,
            'audit_matrix': audit_matrix
        }


def main():
    """主函数"""
    print("🚀 企业政策预审系统测试工具")
    print("="*60)
    print("专注测试：企业数据 × 企业政策 预审")
    print("接口：preaudit-single")
    print("输出：详细Excel预审结果矩阵")
    print("="*60)
    
    # 创建测试器
    tester = EnterprisePreAuditTester()
    
    # 运行测试 - 可以调整参数
    results = tester.run_enterprise_preaudit_test(
        policy_limit=50,      # 最多加载8个政策文件
        enterprise_limit=50  # 最多加载50个企业文件
    )
    
    if results:
        print(f"\n🎉 测试成功完成！")
        print(f"Excel报告: {results['excel_file']}")
        print(f"预审通过率: {results['audit_pass_rate']:.1f}%")
    else:
        print(f"\n❌ 测试失败")


if __name__ == "__main__":
    main()