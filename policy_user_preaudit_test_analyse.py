"""
政策预审系统完整测试套件
包含性能测试、精度测试、参数调优测试等
增强版：添加详细预审结果Excel报告生成功能
"""

import requests
import json
import time
import os
import pandas as pd
from typing import List, Dict, Any
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class PolicyPreAuditTester:
    """政策预审系统测试类"""
    
    def __init__(self, base_url: str = "http://127.0.0.1:8082"):
        self.base_url = base_url
        self.test_results = {
            'api_tests': [],
            'performance_tests': [],
            'accuracy_tests': [],
            'parameter_tuning': []
        }
        self.audit_matrix = []  # 存储所有预审结果
        
    def load_test_data(self, policy_folder: str = "policy_new", user_folder: str = "user_dataset", 
                       policy_limit: int = 8, user_limit: int = 50):
        """加载测试数据"""
        logger.info(f"开始加载测试数据 - 政策文件夹:{policy_folder}, 用户文件夹:{user_folder}")
        
        # 加载政策数据
        policies = []
        if os.path.exists(policy_folder):
            policy_files = sorted([f for f in os.listdir(policy_folder) if f.endswith('.json')])
            logger.info(f"发现政策文件: {policy_files}")
            
            for file in policy_files[:policy_limit]:
                try:
                    with open(os.path.join(policy_folder, file), 'r', encoding='utf-8') as f:
                        policy_data = json.load(f)
                        policies.append(policy_data)
                        logger.info(f"✅ 加载政策文件: {file} - {policy_data.get('标题', '无标题')}")
                except Exception as e:
                    logger.warning(f"❌ 加载政策文件 {file} 失败: {e}")
        else:
            logger.error(f"❌ 政策文件夹不存在: {policy_folder}")
            return [], []
        
        # 加载用户数据
        users = []
        if os.path.exists(user_folder):
            user_files = sorted([f for f in os.listdir(user_folder) if f.endswith('.json')])
            logger.info(f"发现用户文件: {user_files}")
            
            for file in user_files[:user_limit]:
                try:
                    with open(os.path.join(user_folder, file), 'r', encoding='utf-8') as f:
                        user_data = json.load(f)
                        users.append(user_data)
                        logger.info(f"✅ 加载用户文件: {file} - 用户ID: {user_data.get('用户ID', '无ID')}")
                except Exception as e:
                    logger.warning(f"❌ 加载用户文件 {file} 失败: {e}")
        else:
            logger.error(f"❌ 用户文件夹不存在: {user_folder}")
            return [], []
        
        if not policies:
            logger.error("❌ 没有成功加载任何政策数据")
            return [], []
        
        if not users:
            logger.error("❌ 没有成功加载任何用户数据")
            return [], []
        
        logger.info(f"✅ 数据加载完成 - 政策数量:{len(policies)}, 用户数量:{len(users)}")
        return policies, users
    
    def test_health_check(self):
        """测试健康检查接口"""
        logger.info("🩺 开始健康检查测试")
        start_time = time.time()
        
        try:
            response = requests.get(f"{self.base_url}/health", timeout=10)
            end_time = time.time()
            
            test_result = {
                'test_name': '健康检查',
                'endpoint': '/health',
                'method': 'GET',
                'status_code': response.status_code,
                'response_time': round((end_time - start_time) * 1000, 2),
                'success': response.status_code == 200,
                'response_data': response.json() if response.status_code == 200 else None,
                'timestamp': datetime.now().isoformat()
            }
            
            self.test_results['api_tests'].append(test_result)
            logger.info(f"✅ 健康检查成功 - 响应时间: {test_result['response_time']}ms")
            return test_result
            
        except Exception as e:
            test_result = {
                'test_name': '健康检查',
                'endpoint': '/health',
                'method': 'GET',
                'status_code': 0,
                'response_time': 0,
                'success': False,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
            self.test_results['api_tests'].append(test_result)
            logger.error(f"❌ 健康检查失败: {e}")
            return test_result
    
    def test_preaudit_single(self, user_data: Dict, policy_data: Dict):
        """测试单个政策预审接口"""
        start_time = time.time()
        
        try:
            payload = {
                "user": user_data,
                "policy": policy_data
            }
            
            response = requests.post(
                f"{self.base_url}/preaudit-single",
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=30
            )
            end_time = time.time()
            
            test_result = {
                'test_name': '单个政策预审',
                'endpoint': '/preaudit-single',
                'method': 'POST',
                'user_id': user_data.get('用户ID'),
                'policy_id': policy_data.get('政策编号'),
                'status_code': response.status_code,
                'response_time': round((end_time - start_time) * 1000, 2),
                'success': response.status_code == 200,
                'response_data': response.json() if response.status_code == 200 else None,
                'timestamp': datetime.now().isoformat()
            }
            
            if response.status_code == 200:
                result_data = response.json()
                test_result['audit_result'] = result_data.get('result', 0)  # 0或1
            
            return test_result
            
        except Exception as e:
            test_result = {
                'test_name': '单个政策预审',
                'endpoint': '/preaudit-single',
                'method': 'POST',
                'user_id': user_data.get('用户ID'),
                'policy_id': policy_data.get('政策编号'),
                'status_code': 0,
                'response_time': 0,
                'success': False,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
            return test_result
    
    def generate_audit_matrix(self, users: List[Dict], policies: List[Dict]):
        """生成用户-政策预审结果矩阵"""
        logger.info(f"🔍 开始生成预审结果矩阵 - {len(users)}个用户 × {len(policies)}个政策")
        
        audit_matrix = []
        total_tests = len(users) * len(policies)
        current_test = 0
        
        for user in users:
            user_id = user.get('用户ID')
            user_audits = {
                'user_id': user_id,
                'user_info': user,
                'policy_audits': []
            }
            
            logger.info(f"测试用户 {user_id} 的政策预审结果...")
            
            for policy in policies:
                current_test += 1
                policy_id = policy.get('政策编号')
                
                # 获取预审结果
                result = self.test_preaudit_single(user, policy)
                audit_result = result.get('audit_result', 0) if result.get('success', False) else 0
                
                user_audits['policy_audits'].append({
                    'policy_id': policy_id,
                    'policy_info': policy,
                    'audit_result': audit_result,  # 0或1
                    'success': result.get('success', False)
                })
                
                status = "通过" if audit_result == 1 else "不通过"
                logger.info(f"  {policy_id}: {status} ({audit_result}) ({current_test}/{total_tests})")
                time.sleep(0.1)  # 避免请求过于频繁
            
            audit_matrix.append(user_audits)
        
        self.audit_matrix = audit_matrix
        logger.info(f"✅ 预审结果矩阵生成完成")
        return audit_matrix
    
    def format_policy_conditions(self, policy: Dict) -> str:
        """格式化政策条件"""
        def parse_rules(rules, parent_logic="AND"):
            if not isinstance(rules, list):
                return []
            
            condition_groups = []
            
            for rule in rules:
                if not isinstance(rule, dict):
                    continue
                
                if '规则' in rule and isinstance(rule['规则'], list):
                    nested_logic = rule.get('逻辑', 'AND')
                    nested_conditions = parse_rules(rule['规则'], nested_logic)
                    if nested_conditions:
                        if nested_logic.lower() == 'or':
                            condition_groups.append(f"({' 或 '.join(nested_conditions)})")
                        else:
                            condition_groups.append(f"({' 且 '.join(nested_conditions)})")
                else:
                    field = rule.get('字段', '')
                    value = rule.get('值', '')
                    operator = rule.get('操作符', '')
                    desc = rule.get('描述', '')
                    
                    if field and value is not None:
                        if operator == '<=':
                            condition_str = f"{field}≤{value}"
                        elif operator == '>=':
                            condition_str = f"{field}≥{value}"
                        elif operator == '=':
                            condition_str = f"{field}={value}"
                        elif operator == '!=':
                            condition_str = f"{field}≠{value}"
                        elif operator == '<':
                            condition_str = f"{field}<{value}"
                        elif operator == '>':
                            condition_str = f"{field}>{value}"
                        elif operator == 'between' and isinstance(value, list) and len(value) == 2:
                            condition_str = f"{field}在{value[0]}-{value[1]}之间"
                        elif operator == 'in' and isinstance(value, list):
                            condition_str = f"{field} in [{','.join(map(str, value))}]"
                        else:
                            condition_str = f"{field}:{value}"
                        
                        if desc:
                            condition_str = f"{desc}({condition_str})"
                        
                        condition_groups.append(condition_str)
            
            return condition_groups
        
        conditions = policy.get('条件', {})
        rules = conditions.get('规则', [])
        main_logic = conditions.get('逻辑', 'AND')
        
        if not rules:
            return "无明确条件"
        
        all_conditions = parse_rules(rules, main_logic)
        
        if not all_conditions:
            return "条件解析失败"
        
        if main_logic.lower() == 'or':
            return " 或 ".join(all_conditions)
        else:
            return " 且 ".join(all_conditions)
    
    def format_user_info(self, user: Dict) -> str:
        """格式化用户信息"""
        info_parts = []
        key_fields = [
            ('用户ID', ''), ('最高学历', ''), ('年龄', '岁'), ('毕业时间', '年'),
            ('专业', ''), ('就业类型', ''), ('缴纳社保', ''), ('养老保险', ''),
            ('征地人员', ''), ('困难人员', '')
        ]
        
        for field, unit in key_fields:
            if field in user and user[field] is not None:
                value = user[field]
                if unit:
                    info_parts.append(f"{field}:{value}{unit}")
                else:
                    info_parts.append(f"{field}:{value}")
        
        return " | ".join(info_parts)
    
    def export_to_excel(self, filename: str = 'policy_preaudit_matrix.xlsx'):
        """导出预审结果矩阵到Excel"""
        if not self.audit_matrix:
            logger.error("❌ 没有预审数据可以导出")
            return None
        
        logger.info(f"📊 开始导出Excel报告: {filename}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "政策预审结果矩阵"
        
        policies = self.audit_matrix[0]['policy_audits']
        users = self.audit_matrix
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        policy_font = Font(bold=True, color="FFFFFF")
        policy_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        user_font = Font(bold=True)
        user_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 构建表格
        ws.merge_cells('A1:B1')
        ws['A1'] = "用户信息 / 政策编号"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].border = thin_border
        
        # 政策编号行
        col = 3
        for policy_audit in policies:
            ws.cell(row=1, column=col, value=policy_audit['policy_id'])
            ws.cell(row=1, column=col).font = policy_font
            ws.cell(row=1, column=col).fill = policy_fill
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=col).border = thin_border
            col += 1
        
        # 政策标题行
        ws.merge_cells('A2:B2')
        ws['A2'] = "政策标题"
        ws['A2'].font = header_font
        ws['A2'].fill = header_fill
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].border = thin_border
        
        col = 3
        for policy_audit in policies:
            policy_title = policy_audit['policy_info'].get('标题', '无标题')
            ws.cell(row=2, column=col, value=policy_title)
            ws.cell(row=2, column=col).font = Font(size=10, bold=True)
            ws.cell(row=2, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=2, column=col).border = thin_border
            col += 1
        
        # 政策条件行
        ws.merge_cells('A3:B3')
        ws['A3'] = "政策条件"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A3'].border = thin_border
        
        col = 3
        for policy_audit in policies:
            conditions = self.format_policy_conditions(policy_audit['policy_info'])
            ws.cell(row=3, column=col, value=conditions)
            ws.cell(row=3, column=col).font = Font(size=9)
            ws.cell(row=3, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.cell(row=3, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.cell(row=3, column=col).border = thin_border
            col += 1
        
        # 用户数据和预审结果
        row = 4
        for user_audit in users:
            user_id = user_audit['user_id']
            user_desc = self.format_user_info(user_audit['user_info'])
            
            # A列：用户ID
            ws.cell(row=row, column=1, value=user_id)
            ws.cell(row=row, column=1).font = user_font
            ws.cell(row=row, column=1).fill = user_fill
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=1).border = thin_border
            
            # B列：用户信息
            ws.cell(row=row, column=2, value=user_desc)
            ws.cell(row=row, column=2).font = Font(size=9)
            ws.cell(row=row, column=2).fill = user_fill
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(row=row, column=2).border = thin_border
            
            # 预审结果
            col = 3
            for policy_audit in user_audit['policy_audits']:
                audit_result = policy_audit['audit_result']
                
                # 直接显示0或1
                ws.cell(row=row, column=col, value=audit_result)
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).font = Font(bold=True)
                
                # 淡色编码
                if audit_result == 1:
                    # 淡绿色
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    ws.cell(row=row, column=col).font = Font(color="2E7D32", bold=True)
                else:
                    # 淡红色
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    ws.cell(row=row, column=col).font = Font(color="C62828", bold=True)
                
                col += 1
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        for col in range(3, 3 + len(policies)):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # 调整行高
        ws.row_dimensions[2].height = 60
        ws.row_dimensions[3].height = 80
        for row_num in range(4, 4 + len(users)):
            ws.row_dimensions[row_num].height = 60
        
        # 添加统计分析工作表
        stats_ws = wb.create_sheet("统计分析")
        self._add_statistics_sheet(stats_ws)
        
        wb.save(filename)
        logger.info(f"✅ Excel报告已保存到: {filename}")
        return filename
    
    def _add_statistics_sheet(self, ws):
        """添加统计分析工作表"""
        ws['A1'] = "政策预审结果统计分析"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # 用户统计
        row = 3
        ws[f'A{row}'] = "用户预审结果统计:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ["用户ID", "通过政策数", "总政策数", "通过率(%)"]
        for i, header in enumerate(headers):
            ws.cell(row=row, column=i+1, value=header)
            ws.cell(row=row, column=i+1).font = Font(bold=True)
            ws.cell(row=row, column=i+1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row += 1
        for user_audit in self.audit_matrix:
            user_id = user_audit['user_id']
            results = [pa['audit_result'] for pa in user_audit['policy_audits']]
            passed_count = sum(results)
            total_count = len(results)
            pass_rate = round(passed_count / total_count * 100, 1) if total_count > 0 else 0
            
            ws.cell(row=row, column=1, value=user_id)
            ws.cell(row=row, column=2, value=passed_count)
            ws.cell(row=row, column=3, value=total_count)
            ws.cell(row=row, column=4, value=pass_rate)
            row += 1
        
        # 政策统计
        row += 2
        ws[f'A{row}'] = "政策预审结果统计:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ["政策编号", "政策标题", "通过用户数", "通过率(%)"]
        for i, header in enumerate(headers):
            ws.cell(row=row, column=i+1, value=header)
            ws.cell(row=row, column=i+1).font = Font(bold=True)
            ws.cell(row=row, column=i+1).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        row += 1
        policy_stats = {}
        for user_audit in self.audit_matrix:
            for policy_audit in user_audit['policy_audits']:
                policy_id = policy_audit['policy_id']
                if policy_id not in policy_stats:
                    policy_stats[policy_id] = {
                        'title': policy_audit['policy_info'].get('标题', ''),
                        'results': []
                    }
                policy_stats[policy_id]['results'].append(policy_audit['audit_result'])
        
        for policy_id, stats in policy_stats.items():
            results = stats['results']
            passed_count = sum(results)
            total_count = len(results)
            pass_rate = round(passed_count / total_count * 100, 1) if total_count > 0 else 0
            
            ws.cell(row=row, column=1, value=policy_id)
            ws.cell(row=row, column=2, value=stats['title'])
            ws.cell(row=row, column=3, value=passed_count)
            ws.cell(row=row, column=4, value=pass_rate)
            row += 1
        
        # 调整列宽
        for col, width in [('A', 15), ('B', 30), ('C', 15), ('D', 15)]:
            ws.column_dimensions[col].width = width
    
    def run_full_test_suite(self, policy_limit: int = 8, user_limit: int = 50):
        """运行完整测试套件"""
        logger.info("🚀 开始运行完整预审测试套件")
        test_start_time = time.time()
        
        # 加载测试数据
        policies, users = self.load_test_data(policy_limit=policy_limit, user_limit=user_limit)
        
        if not policies or not users:
            logger.error("❌ 测试数据加载失败")
            return None
        
        # 1. 健康检查测试
        logger.info("\n" + "="*50)
        logger.info("1. 健康检查测试")
        logger.info("="*50)
        self.test_health_check()
        
        # 2. 生成完整预审结果矩阵
        logger.info("\n" + "="*50)
        logger.info("2. 生成完整预审结果矩阵")
        logger.info("="*50)
        self.generate_audit_matrix(users, policies)
        
        # 3. 导出Excel报告
        logger.info("\n" + "="*50)
        logger.info("3. 导出Excel报告")
        logger.info("="*50)
        excel_filename = f"policy_preaudit_matrix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.export_to_excel(excel_filename)
        
        test_end_time = time.time()
        total_test_time = test_end_time - test_start_time
        logger.info(f"\n✅ 完整预审测试套件完成 - 总耗时: {total_test_time:.2f}秒")
        
        return {
            'total_test_time': total_test_time,
            'test_data_summary': {
                'policies_count': len(policies),
                'users_count': len(users)
            },
            'excel_file': excel_filename,
            'test_results': self.test_results
        }
    
    def generate_test_report(self, output_file: str = 'preaudit_test_report.json'):
        """生成详细的测试报告"""
        logger.info(f"📝 生成测试报告: {output_file}")
        
        api_tests = self.test_results['api_tests']
        
        # 计算预审统计
        audit_summary = {}
        if self.audit_matrix:
            all_results = []
            for user_audit in self.audit_matrix:
                results = [pa['audit_result'] for pa in user_audit['policy_audits']]
                all_results.extend(results)
            
            audit_summary = {
                'total_audits': len(all_results),
                'total_passed': sum(all_results),
                'total_failed': len(all_results) - sum(all_results),
                'overall_pass_rate': round(sum(all_results) / len(all_results) * 100, 2) if all_results else 0
            }
        
        report = {
            'test_summary': {
                'test_time': datetime.now().isoformat(),
                'total_api_tests': len(api_tests),
                'successful_api_tests': len([t for t in api_tests if t['success']]),
                'audit_matrix_generated': len(self.audit_matrix) > 0
            },
            'audit_summary': audit_summary,
            'detailed_results': self.test_results
        }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        logger.info(f"✅ 测试报告已保存到: {output_file}")
        return report

def main():
    """主函数 - 运行预审测试"""
    print("🚀 政策预审系统测试开始")
    print("="*80)
    
    # 创建测试器实例
    tester = PolicyPreAuditTester()
    
    # 运行完整测试套件
    results = tester.run_full_test_suite(policy_limit=8, user_limit=50)
    
    if results:
        # 生成测试报告
        report_filename = f"policy_preaudit_test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        report = tester.generate_test_report(report_filename)
        
        print("\n" + "="*80)
        print("📊 测试总结")
        print("="*80)
        print(f"测试数据: {results['test_data_summary']['policies_count']} 个政策, {results['test_data_summary']['users_count']} 个用户")
        print(f"总测试时间: {results['total_test_time']:.2f} 秒")
        print(f"Excel报告: {results['excel_file']}")
        
        # 显示预审统计
        if 'audit_summary' in report:
            audit_summary = report['audit_summary']
            print(f"\n📈 预审结果统计:")
            print(f"  总预审测试: {audit_summary.get('total_audits', 0)} 次")
            print(f"  通过总数: {audit_summary.get('total_passed', 0)}")
            print(f"  不通过总数: {audit_summary.get('total_failed', 0)}")
            print(f"  总体通过率: {audit_summary.get('overall_pass_rate', 0):.1f}%")
        
        if report['test_summary']['successful_api_tests'] > 0:
            success_rate = len([t for t in tester.test_results['api_tests'] if t['success']]) / len(tester.test_results['api_tests']) * 100
            print(f"\n✅ API测试成功率: {success_rate:.1f}%")
        
        print(f"\n📝 详细报告已保存到: {report_filename}")
        print(f"📊 Excel预审结果矩阵已保存到: {results['excel_file']}")

        print("="*80)
        print("🎉 测试完成！请查看Excel文件获取详细的用户-政策预审结果矩阵")
        print("   绿色表示通过(1)，红色表示不通过(0)")
    else:
        print("❌ 测试失败")

if __name__ == "__main__":
    main()