"""
政策推荐系统完整测试套件
包含性能测试、精度测试、参数调优测试等
增强版：添加详细匹配度Excel报告生成功能
"""

import requests
import json
import time
import os
import pandas as pd
from typing import List, Dict, Any
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class PolicyRecommendationTester:
    """政策推荐系统测试类"""
    
    def __init__(self, base_url: str = "http://10.1.50.96:8081"):
        self.base_url = base_url
        self.test_results = {
            'api_tests': [],
            'performance_tests': [],
            'accuracy_tests': [],
            'parameter_tuning': []
        }
        self.match_matrix = []  # 存储所有匹配度结果
        
    def load_test_data(self, policy_folder: str = "policy_new", user_folder: str = "user_dataset", 
                       policy_limit: int = 8, user_limit: int = 50):
        """加载测试数据"""
        logger.info(f"开始加载测试数据 - 政策文件夹:{policy_folder}, 用户文件夹:{user_folder}")
        
        # 加载政策数据
        policies = []
        if os.path.exists(policy_folder):
            policy_files = sorted([f for f in os.listdir(policy_folder) if f.endswith('.json')])
            logger.info(f"发现政策文件: {policy_files}")
            
            for file in policy_files[:policy_limit]:
                try:
                    with open(os.path.join(policy_folder, file), 'r', encoding='utf-8') as f:
                        policy_data = json.load(f)
                        policies.append(policy_data)
                        logger.info(f"✅ 加载政策文件: {file} - {policy_data.get('标题', '无标题')}")
                except Exception as e:
                    logger.warning(f"❌ 加载政策文件 {file} 失败: {e}")
        else:
            logger.error(f"❌ 政策文件夹不存在: {policy_folder}")
            return [], []
        
        # 加载用户数据
        users = []
        if os.path.exists(user_folder):
            user_files = sorted([f for f in os.listdir(user_folder) if f.endswith('.json')])
            logger.info(f"发现用户文件: {user_files}")
            
            for file in user_files[:user_limit]:
                try:
                    with open(os.path.join(user_folder, file), 'r', encoding='utf-8') as f:
                        user_data = json.load(f)
                        users.append(user_data)
                        logger.info(f"✅ 加载用户文件: {file} - 用户ID: {user_data.get('用户ID', '无ID')}")
                except Exception as e:
                    logger.warning(f"❌ 加载用户文件 {file} 失败: {e}")
        else:
            logger.error(f"❌ 用户文件夹不存在: {user_folder}")
            return [], []
        
        if not policies:
            logger.error("❌ 没有成功加载任何政策数据")
            return [], []
        
        if not users:
            logger.error("❌ 没有成功加载任何用户数据")
            return [], []
        
        logger.info(f"✅ 数据加载完成 - 政策数量:{len(policies)}, 用户数量:{len(users)}")
        return policies, users
    
    def test_health_check(self):
        """测试健康检查接口"""
        logger.info("🩺 开始健康检查测试")
        start_time = time.time()
        
        try:
            response = requests.get(f"{self.base_url}/health", timeout=10)
            end_time = time.time()
            
            test_result = {
                'test_name': '健康检查',
                'endpoint': '/health',
                'method': 'GET',
                'status_code': response.status_code,
                'response_time': round((end_time - start_time) * 1000, 2),
                'success': response.status_code == 200,
                'response_data': response.json() if response.status_code == 200 else None,
                'timestamp': datetime.now().isoformat()
            }
            
            self.test_results['api_tests'].append(test_result)
            logger.info(f"✅ 健康检查成功 - 响应时间: {test_result['response_time']}ms")
            return test_result
            
        except Exception as e:
            test_result = {
                'test_name': '健康检查',
                'endpoint': '/health',
                'method': 'GET',
                'status_code': 0,
                'response_time': 0,
                'success': False,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
            self.test_results['api_tests'].append(test_result)
            logger.error(f"❌ 健康检查失败: {e}")
            return test_result
    
    def test_recommend_single(self, user_data: Dict, policy_data: Dict):
        """测试单个政策推荐接口"""
        start_time = time.time()
        
        try:
            payload = {
                "user": user_data,
                "policy": policy_data
            }
            
            response = requests.post(
                f"{self.base_url}/recommend-single",
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=30
            )
            end_time = time.time()
            
            test_result = {
                'test_name': '单个政策推荐',
                'endpoint': '/recommend-single',
                'method': 'POST',
                'user_id': user_data.get('用户ID'),
                'policy_id': policy_data.get('政策编号'),
                'status_code': response.status_code,
                'response_time': round((end_time - start_time) * 1000, 2),
                'success': response.status_code == 200,
                'response_data': response.json() if response.status_code == 200 else None,
                'timestamp': datetime.now().isoformat()
            }
            
            if response.status_code == 200:
                result_data = response.json()
                test_result['match_score'] = result_data.get('result', 0)
            
            return test_result
            
        except Exception as e:
            test_result = {
                'test_name': '单个政策推荐',
                'endpoint': '/recommend-single',
                'method': 'POST',
                'user_id': user_data.get('用户ID'),
                'policy_id': policy_data.get('政策编号'),
                'status_code': 0,
                'response_time': 0,
                'success': False,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
            return test_result
    
    def generate_match_matrix(self, users: List[Dict], policies: List[Dict]):
        """生成用户-政策匹配度矩阵"""
        logger.info(f"🔍 开始生成匹配度矩阵 - {len(users)}个用户 × {len(policies)}个政策")
        
        match_matrix = []
        total_tests = len(users) * len(policies)
        current_test = 0
        
        for user in users:
            user_id = user.get('用户ID')
            user_matches = {
                'user_id': user_id,
                'user_info': user,
                'policy_matches': []
            }
            
            logger.info(f"测试用户 {user_id} 的政策匹配度...")
            
            for policy in policies:
                current_test += 1
                policy_id = policy.get('政策编号')
                
                # 获取匹配度
                result = self.test_recommend_single(user, policy)
                match_score = result.get('match_score', 0) if result.get('success', False) else 0
                
                user_matches['policy_matches'].append({
                    'policy_id': policy_id,
                    'policy_info': policy,
                    'match_score': match_score,
                    'success': result.get('success', False)
                })
                
                logger.info(f"  {policy_id}: {match_score:.4f} ({current_test}/{total_tests})")
                time.sleep(0.1)  # 避免请求过于频繁
            
            match_matrix.append(user_matches)
        
        self.match_matrix = match_matrix
        logger.info(f"✅ 匹配度矩阵生成完成")
        return match_matrix
    
    def format_policy_conditions(self, policy: Dict) -> str:
        """简化格式化政策条件 - 支持嵌套规则结构"""
        
        def parse_rules(rules, parent_logic="AND"):
            """递归解析规则，支持嵌套结构"""
            if not isinstance(rules, list):
                return []
            
            condition_groups = []
            
            for rule in rules:
                if not isinstance(rule, dict):
                    continue
                
                # 如果这个规则包含嵌套的规则结构
                if '规则' in rule and isinstance(rule['规则'], list):
                    # 递归处理嵌套规则
                    nested_logic = rule.get('逻辑', 'AND')
                    nested_conditions = parse_rules(rule['规则'], nested_logic)
                    if nested_conditions:
                        if nested_logic == 'OR':
                            condition_groups.append(f"({' 或 '.join(nested_conditions)})")
                        else:
                            condition_groups.append(f"({' 且 '.join(nested_conditions)})")
                else:
                    # 处理单个条件
                    field = rule.get('字段', '')
                    value = rule.get('值', '')
                    operator = rule.get('操作符', '')
                    desc = rule.get('描述', '')
                    
                    if field and value is not None:
                        condition_str = ""
                        
                        if operator == '<=':
                            condition_str = f"{field}≤{value}"
                        elif operator == '>=':
                            condition_str = f"{field}≥{value}"
                        elif operator == '=':
                            condition_str = f"{field}={value}"
                        elif operator == '!=':
                            condition_str = f"{field}≠{value}"
                        elif operator == 'between' and isinstance(value, list) and len(value) == 2:
                            condition_str = f"{field}在{value[0]}-{value[1]}之间"
                        elif operator == 'in' and isinstance(value, list):
                            condition_str = f"{field} in [{','.join(map(str, value))}]"
                        else:
                            condition_str = f"{field}:{value}"
                        
                        # 如果有描述，优先使用描述
                        if desc:
                            condition_str = f"{desc}({condition_str})"
                        
                        condition_groups.append(condition_str)
            
            return condition_groups
        
        # 提取主要条件字段
        conditions = policy.get('条件', {})
        rules = conditions.get('规则', [])
        main_logic = conditions.get('逻辑', 'AND')
        
        # 如果没有条件规则，尝试其他可能的字段名
        if not rules:
            for possible_key in ['条件规则', 'rules', '规则列表', 'conditions']:
                if possible_key in policy:
                    rules = policy[possible_key]
                    if isinstance(rules, list):
                        break
                    elif isinstance(rules, dict) and '规则' in rules:
                        rules = rules['规则']
                        break
        
        # 如果仍然没有规则，记录详细信息
        if not rules:
            logger.warning(f"政策 {policy.get('政策编号', 'Unknown')} 未找到条件规则")
            logger.warning(f"政策结构: {json.dumps(policy, ensure_ascii=False, indent=2)}")
            return "无明确条件"
        
        # 解析所有条件
        all_conditions = parse_rules(rules, main_logic)
        
        if not all_conditions:
            return "条件解析失败"
        
        # 根据主逻辑组合条件
        if main_logic == 'OR':
            return " 或 ".join(all_conditions)
        else:
            return " 且 ".join(all_conditions)
    
    def format_user_info(self, user: Dict) -> str:
        """简化格式化用户信息"""
        info_parts = []
        
        # 按重要性排序的关键字段
        key_fields = [
            ('用户ID', ''),
            ('最高学历', ''),
            ('年龄', '岁'),
            ('毕业时间', '年'),
            ('专业', ''),
            ('就业类型', ''),
            ('缴纳社保', ''),
            ('养老保险', ''),
            ('征地人员', ''),
            ('困难人员', '')
        ]
        
        for field, unit in key_fields:
            if field in user and user[field] is not None:
                value = user[field]
                if unit:
                    info_parts.append(f"{field}:{value}{unit}")
                else:
                    info_parts.append(f"{field}:{value}")
        
        return " | ".join(info_parts)
    
    def export_to_excel(self, filename: str = 'policy_match_matrix.xlsx'):
        """导出匹配度矩阵到Excel"""
        if not self.match_matrix:
            logger.error("❌ 没有匹配度数据可以导出")
            return None
        
        logger.info(f"📊 开始导出Excel报告: {filename}")
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "政策匹配度矩阵"
        
        # 获取所有政策信息
        policies = self.match_matrix[0]['policy_matches']
        users = self.match_matrix
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        policy_font = Font(bold=True, color="FFFFFF")
        policy_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        user_font = Font(bold=True)
        user_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 第一行：政策编号
        ws.merge_cells('A1:B1')
        ws['A1'] = "用户信息 / 政策编号"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].border = thin_border
        
        col = 3  # 从C列开始
        for policy_match in policies:
            policy_id = policy_match['policy_id']
            ws.cell(row=1, column=col, value=policy_id)
            ws.cell(row=1, column=col).font = policy_font
            ws.cell(row=1, column=col).fill = policy_fill
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=col).border = thin_border
            col += 1
        
        # 第二行：政策标题
        ws.merge_cells('A2:B2')
        ws['A2'] = "政策标题"
        ws['A2'].font = header_font
        ws['A2'].fill = header_fill
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].border = thin_border
        
        col = 3
        for policy_match in policies:
            policy_info = policy_match['policy_info']
            policy_title = policy_info.get('标题', '无标题')
            
            ws.cell(row=2, column=col, value=policy_title)
            ws.cell(row=2, column=col).font = Font(size=10, bold=True)
            ws.cell(row=2, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=2, column=col).border = thin_border
            col += 1
        
        # 第三行：政策条件
        ws.merge_cells('A3:B3')
        ws['A3'] = "政策条件"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A3'].border = thin_border
        
        col = 3
        for policy_match in policies:
            policy_info = policy_match['policy_info']
            conditions = self.format_policy_conditions(policy_info)
            
            ws.cell(row=3, column=col, value=conditions)
            ws.cell(row=3, column=col).font = Font(size=9)
            ws.cell(row=3, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.cell(row=3, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.cell(row=3, column=col).border = thin_border
            col += 1
        
        # 从第四行开始：用户数据和匹配度
        row = 4
        for user_match in users:
            user_id = user_match['user_id']
            user_info = user_match['user_info']
            user_desc = self.format_user_info(user_info)
            
            # A列：用户ID
            ws.cell(row=row, column=1, value=user_id)
            ws.cell(row=row, column=1).font = user_font
            ws.cell(row=row, column=1).fill = user_fill
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=1).border = thin_border
            
            # B列：用户信息
            ws.cell(row=row, column=2, value=user_desc)
            ws.cell(row=row, column=2).font = Font(size=9)
            ws.cell(row=row, column=2).fill = user_fill
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(row=row, column=2).border = thin_border
            
            # 从C列开始：匹配度分数
            col = 3
            for policy_match in user_match['policy_matches']:
                match_score = policy_match['match_score']
                
                ws.cell(row=row, column=col, value=match_score)
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).number_format = '0.00'
                
                # 根据匹配度设置颜色
                if match_score >= 0.8:
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                elif match_score >= 0.6:
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif match_score >= 0.3:
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                else:
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    ws.cell(row=row, column=col).font = Font(color="FFFFFF")
                
                col += 1
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        for col in range(3, 3 + len(policies)):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # 调整行高
        ws.row_dimensions[2].height = 60  # 政策标题行
        ws.row_dimensions[3].height = 80  # 政策条件行
        for row_num in range(4, 4 + len(users)):
            ws.row_dimensions[row_num].height = 60  # 用户信息行
        
        # 添加统计信息工作表
        stats_ws = wb.create_sheet("统计分析")
        self._add_statistics_sheet(stats_ws, users, policies)
        
        # 保存文件
        wb.save(filename)
        logger.info(f"✅ Excel报告已保存到: {filename}")
        
        return filename
    
    def _add_statistics_sheet(self, ws, users, policies):
        """添加统计分析工作表"""
        # 标题
        ws['A1'] = "政策匹配度统计分析"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # 用户统计
        ws[f'A{row}'] = "用户匹配度统计:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "用户ID"
        ws[f'B{row}'] = "平均匹配度"
        ws[f'C{row}'] = "最高匹配度"
        ws[f'D{row}'] = "匹配政策数(>0.5)"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row += 1
        
        for user_match in self.match_matrix:
            user_id = user_match['user_id']
            scores = [pm['match_score'] for pm in user_match['policy_matches']]
            
            avg_score = np.mean(scores)
            max_score = max(scores)
            high_match_count = sum(1 for score in scores if score > 0.5)
            
            ws[f'A{row}'] = user_id
            ws[f'B{row}'] = round(avg_score, 4)
            ws[f'C{row}'] = round(max_score, 4)
            ws[f'D{row}'] = high_match_count
            row += 1
        
        row += 2
        
        # 政策统计
        ws[f'A{row}'] = "政策匹配度统计:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "政策编号"
        ws[f'B{row}'] = "政策标题"
        ws[f'C{row}'] = "平均匹配度"
        ws[f'D{row}'] = "匹配用户数(>0.5)"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        row += 1
        
        # 按政策统计
        policy_stats = {}
        for user_match in self.match_matrix:
            for policy_match in user_match['policy_matches']:
                policy_id = policy_match['policy_id']
                if policy_id not in policy_stats:
                    policy_stats[policy_id] = {
                        'title': policy_match['policy_info'].get('标题', ''),
                        'scores': []
                    }
                policy_stats[policy_id]['scores'].append(policy_match['match_score'])
        
        for policy_id, stats in policy_stats.items():
            scores = stats['scores']
            avg_score = np.mean(scores)
            high_match_count = sum(1 for score in scores if score > 0.5)
            
            ws[f'A{row}'] = policy_id
            ws[f'B{row}'] = stats['title']
            ws[f'C{row}'] = round(avg_score, 4)
            ws[f'D{row}'] = high_match_count
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
    
    def test_recommend_batch(self, user_data: Dict, policies: List[Dict]):
        """测试批量推荐接口（单用户多政策）"""
        logger.info(f"📊 批量推荐测试 - 用户:{user_data.get('用户ID')}, 政策数量:{len(policies)}")
        start_time = time.time()
        
        try:
            payload = {
                "user": user_data,
                "policies": policies
            }
            
            response = requests.post(
                f"{self.base_url}/recommend",
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=60
            )
            end_time = time.time()
            
            test_result = {
                'test_name': '批量推荐（单用户多政策）',
                'endpoint': '/recommend',
                'method': 'POST',
                'user_id': user_data.get('用户ID'),
                'policy_count': len(policies),
                'status_code': response.status_code,
                'response_time': round((end_time - start_time) * 1000, 2),
                'success': response.status_code == 200,
                'timestamp': datetime.now().isoformat()
            }
            
            if response.status_code == 200:
                result_data = response.json()
                scores = result_data.get('result', [])
                test_result['match_scores'] = scores
                test_result['avg_score'] = round(np.mean(scores), 2) if scores else 0
                test_result['max_score'] = max(scores) if scores else 0
                test_result['min_score'] = min(scores) if scores else 0
                
                # 创建详细的政策匹配结果
                policy_results = []
                for i, score in enumerate(scores):
                    if i < len(policies):
                        policy_results.append({
                            'policy_id': policies[i].get('政策编号'),
                            'policy_title': policies[i].get('标题'),
                            'match_score': score
                        })
                test_result['policy_results'] = policy_results
                
                logger.info(f"✅ 批量推荐成功 - 平均分数: {test_result['avg_score']}")
            else:
                logger.error(f"❌ 批量推荐失败 - 状态码: {response.status_code}")
                test_result['response_data'] = response.text
                
            self.test_results['api_tests'].append(test_result)
            return test_result
            
        except Exception as e:
            test_result = {
                'test_name': '批量推荐（单用户多政策）',
                'endpoint': '/recommend',
                'method': 'POST',
                'user_id': user_data.get('用户ID'),
                'policy_count': len(policies),
                'status_code': 0,
                'response_time': 0,
                'success': False,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
            self.test_results['api_tests'].append(test_result)
            logger.error(f"❌ 批量推荐异常: {e}")
            return test_result
    
    def test_batch_recommend_multi_user(self, users: List[Dict], policies: List[Dict]):
        """测试多用户批量推荐接口"""
        logger.info(f"👥 多用户批量推荐测试 - 用户数量:{len(users)}, 政策数量:{len(policies)}")
        start_time = time.time()
        
        try:
            payload = {
                "users": users,
                "policies": policies
            }
            
            response = requests.post(
                f"{self.base_url}/batch-recommend",
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=120
            )
            end_time = time.time()
            
            test_result = {
                'test_name': '多用户批量推荐',
                'endpoint': '/batch-recommend',
                'method': 'POST',
                'user_count': len(users),
                'policy_count': len(policies),
                'status_code': response.status_code,
                'response_time': round((end_time - start_time) * 1000, 2),
                'success': response.status_code == 200,
                'timestamp': datetime.now().isoformat()
            }
            
            if response.status_code == 200:
                result_data = response.json()
                scores = result_data.get('result', [])
                test_result['total_scores'] = len(scores)
                test_result['expected_scores'] = len(users) * len(policies)
                test_result['scores_complete'] = len(scores) == len(users) * len(policies)
                
                if scores:
                    test_result['avg_score'] = round(np.mean(scores), 2)
                    test_result['score_std'] = round(np.std(scores), 2)
                
                logger.info(f"✅ 多用户批量推荐成功 - 总分数数量: {len(scores)}")
            else:
                logger.error(f"❌ 多用户批量推荐失败 - 状态码: {response.status_code}")
                test_result['response_data'] = response.text
                
            self.test_results['api_tests'].append(test_result)
            return test_result
            
        except Exception as e:
            test_result = {
                'test_name': '多用户批量推荐',
                'endpoint': '/batch-recommend',
                'method': 'POST',
                'user_count': len(users),
                'policy_count': len(policies),
                'status_code': 0,
                'response_time': 0,
                'success': False,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
            self.test_results['api_tests'].append(test_result)
            logger.error(f"❌ 多用户批量推荐异常: {e}")
            return test_result
    
    def performance_test(self, users: List[Dict], policies: List[Dict], 
                        concurrent_requests: int = 5, iterations: int = 10):
        """性能测试"""
        logger.info(f"⚡ 性能测试开始 - 并发数:{concurrent_requests}, 迭代次数:{iterations}")
        
        def single_request():
            """单次请求"""
            user = users[np.random.randint(0, len(users))]
            start_time = time.time()
            
            try:
                payload = {"user": user, "policies": policies}
                response = requests.post(
                    f"{self.base_url}/recommend",
                    json=payload,
                    headers={"Content-Type": "application/json"},
                    timeout=30
                )
                end_time = time.time()
                
                return {
                    'success': response.status_code == 200,
                    'response_time': (end_time - start_time) * 1000,
                    'status_code': response.status_code
                }
            except Exception as e:
                end_time = time.time()
                return {
                    'success': False,
                    'response_time': (end_time - start_time) * 1000,
                    'error': str(e)
                }
        
        # 执行性能测试
        all_results = []
        total_start_time = time.time()
        
        for iteration in range(iterations):
            logger.info(f"执行第 {iteration + 1}/{iterations} 轮测试")
            
            with ThreadPoolExecutor(max_workers=concurrent_requests) as executor:
                futures = [executor.submit(single_request) for _ in range(concurrent_requests)]
                
                for future in as_completed(futures):
                    result = future.result()
                    all_results.append(result)
        
        total_end_time = time.time()
        
        # 统计结果
        successful_requests = [r for r in all_results if r['success']]
        failed_requests = [r for r in all_results if not r['success']]
        
        if successful_requests:
            response_times = [r['response_time'] for r in successful_requests]
            
            performance_result = {
                'test_name': '性能测试',
                'concurrent_requests': concurrent_requests,
                'iterations': iterations,
                'total_requests': len(all_results),
                'successful_requests': len(successful_requests),
                'failed_requests': len(failed_requests),
                'success_rate': round(len(successful_requests) / len(all_results) * 100, 2),
                'avg_response_time': round(np.mean(response_times), 2),
                'min_response_time': round(min(response_times), 2),
                'max_response_time': round(max(response_times), 2),
                'p95_response_time': round(np.percentile(response_times, 95), 2),
                'p99_response_time': round(np.percentile(response_times, 99), 2),
                'std_response_time': round(np.std(response_times), 2),
                'total_test_time': round((total_end_time - total_start_time) * 1000, 2),
                'requests_per_second': round(len(all_results) / (total_end_time - total_start_time), 2),
                'timestamp': datetime.now().isoformat()
            }
        else:
            performance_result = {
                'test_name': '性能测试',
                'concurrent_requests': concurrent_requests,
                'iterations': iterations,
                'total_requests': len(all_results),
                'successful_requests': 0,
                'failed_requests': len(failed_requests),
                'success_rate': 0,
                'error': '所有请求都失败了',
                'timestamp': datetime.now().isoformat()
            }
        
        self.test_results['performance_tests'].append(performance_result)
        logger.info(f"✅ 性能测试完成 - 成功率: {performance_result.get('success_rate', 0)}%")
        return performance_result
    
    def accuracy_test(self, users: List[Dict], policies: List[Dict]):
        """精度测试 - 测试算法的准确性和一致性"""
        logger.info("🎯 精度测试开始")
        
        accuracy_results = []
        
        for user in users:
            user_id = user.get('用户ID')
            logger.info(f"测试用户 {user_id} 的匹配精度")
            
            # 获取批量推荐结果
            batch_result = self.test_recommend_batch(user, policies)
            
            if not batch_result['success']:
                continue
                
            batch_scores = batch_result.get('match_scores', [])
            
            # 获取单个推荐结果进行对比
            single_scores = []
            for policy in policies:
                single_result = self.test_recommend_single(user, policy)
                if single_result['success']:
                    single_scores.append(single_result.get('match_score', 0))
                else:
                    single_scores.append(0)
            
            # 计算一致性
            if len(batch_scores) == len(single_scores):
                consistency_score = np.corrcoef(batch_scores, single_scores)[0, 1] if len(batch_scores) > 1 else 1.0
                mse = np.mean((np.array(batch_scores) - np.array(single_scores)) ** 2)
                mae = np.mean(np.abs(np.array(batch_scores) - np.array(single_scores)))
                
                accuracy_result = {
                    'user_id': user_id,
                    'batch_scores': batch_scores,
                    'single_scores': single_scores,
                    'consistency_correlation': round(consistency_score, 4),
                    'mean_squared_error': round(mse, 4),
                    'mean_absolute_error': round(mae, 4),
                    'max_difference': round(max(np.abs(np.array(batch_scores) - np.array(single_scores))), 4),
                    'identical_results': batch_scores == single_scores
                }
                
                accuracy_results.append(accuracy_result)
                logger.info(f"用户 {user_id} 精度测试完成 - 相关性: {consistency_score:.4f}")
        
        # 计算总体精度指标
        if accuracy_results:
            overall_accuracy = {
                'test_name': '精度测试',
                'tested_users': len(accuracy_results),
                'avg_consistency_correlation': round(np.mean([r['consistency_correlation'] for r in accuracy_results]), 4),
                'avg_mse': round(np.mean([r['mean_squared_error'] for r in accuracy_results]), 4),
                'avg_mae': round(np.mean([r['mean_absolute_error'] for r in accuracy_results]), 4),
                'identical_results_count': sum([1 for r in accuracy_results if r['identical_results']]),
                'identical_results_rate': round(sum([1 for r in accuracy_results if r['identical_results']]) / len(accuracy_results) * 100, 2),
                'detailed_results': accuracy_results,
                'timestamp': datetime.now().isoformat()
            }
            
            self.test_results['accuracy_tests'].append(overall_accuracy)
            logger.info(f"✅ 精度测试完成 - 平均相关性: {overall_accuracy['avg_consistency_correlation']}")
            return overall_accuracy
        else:
            logger.error("❌ 精度测试失败 - 无有效结果")
            return None
    
    def run_full_test_suite(self, policy_limit: int = 8, user_limit: int = 10):
        """运行完整测试套件"""
        logger.info("🚀 开始运行完整测试套件")
        test_start_time = time.time()
        
        # 加载测试数据
        policies, users = self.load_test_data(policy_limit=policy_limit, user_limit=user_limit)
        
        if not policies or not users:
            logger.error("❌ 测试数据加载失败")
            return None
        
        # 1. 健康检查测试
        logger.info("\n" + "="*50)
        logger.info("1. 健康检查测试")
        logger.info("="*50)
        self.test_health_check()
        
        # 2. 生成完整匹配度矩阵（这是核心功能）
        logger.info("\n" + "="*50)
        logger.info("2. 生成完整匹配度矩阵")
        logger.info("="*50)
        self.generate_match_matrix(users, policies)
        
        # 3. 导出Excel报告
        logger.info("\n" + "="*50)
        logger.info("3. 导出Excel报告")
        logger.info("="*50)
        excel_filename = f"policy_match_matrix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.export_to_excel(excel_filename)
        
        # 4. 批量推荐测试（选择性执行）
        logger.info("\n" + "="*50)
        logger.info("4. 批量推荐测试（单用户多政策）")
        logger.info("="*50)
        for i, user in enumerate(users[:3]):  # 只测试前3个用户以节省时间
            self.test_recommend_batch(user, policies)
            if i < 2:
                time.sleep(0.1)
        
        # 5. 多用户批量推荐测试
        logger.info("\n" + "="*50)
        logger.info("5. 多用户批量推荐测试")
        logger.info("="*50)
        self.test_batch_recommend_multi_user(users, policies)
        
        # 6. 性能测试（可选）
        logger.info("\n" + "="*50)
        logger.info("6. 性能测试")
        logger.info("="*50)
        self.performance_test(users[:5], policies, concurrent_requests=3, iterations=3)
        
        test_end_time = time.time()
        
        # 생성测试总结
        total_test_time = test_end_time - test_start_time
        logger.info(f"\n✅ 完整测试套件完成 - 总耗时: {total_test_time:.2f}秒")
        
        return {
            'total_test_time': total_test_time,
            'test_data_summary': {
                'policies_count': len(policies),
                'users_count': len(users)
            },
            'excel_file': excel_filename,
            'test_results': self.test_results
        }
    
    def generate_test_report(self, output_file: str = 'test_report.json'):
        """生成详细的测试报告"""
        logger.info(f"📝 生成测试报告: {output_file}")
        
        # 计算测试统计
        api_tests = self.test_results['api_tests']
        performance_tests = self.test_results['performance_tests']
        accuracy_tests = self.test_results['accuracy_tests']
        
        report = {
            'test_summary': {
                'test_time': datetime.now().isoformat(),
                'total_api_tests': len(api_tests),
                'successful_api_tests': len([t for t in api_tests if t['success']]),
                'failed_api_tests': len([t for t in api_tests if not t['success']]),
                'api_success_rate': round(len([t for t in api_tests if t['success']]) / len(api_tests) * 100, 2) if api_tests else 0,
                'performance_tests_count': len(performance_tests),
                'accuracy_tests_count': len(accuracy_tests),
                'match_matrix_generated': len(self.match_matrix) > 0
            },
            'detailed_results': self.test_results,
            'match_matrix_summary': self._get_match_matrix_summary()
        }
        
        # 保存报告
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        logger.info(f"✅ 测试报告已保存到: {output_file}")
        return report
    
    def _get_match_matrix_summary(self):
        """获取匹配度矩阵的统计摘要"""
        if not self.match_matrix:
            return {}
        
        all_scores = []
        user_stats = []
        
        for user_match in self.match_matrix:
            scores = [pm['match_score'] for pm in user_match['policy_matches']]
            all_scores.extend(scores)
            
            user_stats.append({
                'user_id': user_match['user_id'],
                'avg_score': round(np.mean(scores), 4),
                'max_score': round(max(scores), 4),
                'high_matches': sum(1 for s in scores if s > 0.5)
            })
        
        return {
            'total_matches': len(all_scores),
            'overall_avg_score': round(np.mean(all_scores), 4),
            'overall_std_score': round(np.std(all_scores), 4),
            'high_match_rate': round(sum(1 for s in all_scores if s > 0.5) / len(all_scores) * 100, 2),
            'perfect_match_count': sum(1 for s in all_scores if s >= 1.0),
            'user_statistics': user_stats
        }

def main():
    """主函数 - 运行测试"""
    print("🚀 政策推荐系统测试开始")
    print("="*80)
    
    # 创建测试器实例
    tester = PolicyRecommendationTester()
    
    # 运行完整测试套件 - 8个政策，10个用户
    results = tester.run_full_test_suite(policy_limit=8, user_limit=50)
    
    if results:
        # 生成测试报告
        report_filename = f"policy_recommendation_test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        report = tester.generate_test_report(report_filename)
        
        print("\n" + "="*80)
        print("📊 测试总结")
        print("="*80)
        print(f"测试数据: {results['test_data_summary']['policies_count']} 个政策, {results['test_data_summary']['users_count']} 个用户")
        print(f"总测试时间: {results['total_test_time']:.2f} 秒")
        print(f"Excel报告: {results['excel_file']}")
        
        # 显示匹配度矩阵统计
        if 'match_matrix_summary' in report:
            matrix_summary = report['match_matrix_summary']
            print(f"\n📈 匹配度矩阵统计:")
            print(f"  总匹配测试: {matrix_summary.get('total_matches', 0)} 次")
            print(f"  平均匹配度: {matrix_summary.get('overall_avg_score', 0):.4f}")
            print(f"  高匹配率(>0.5): {matrix_summary.get('high_match_rate', 0):.1f}%")
            print(f"  完美匹配数: {matrix_summary.get('perfect_match_count', 0)}")
        
        if report['test_summary']['successful_api_tests'] > 0:
            print(f"\n✅ API测试成功率: {report['test_summary']['api_success_rate']}%")
            
            # 显示性能测试结果
            if tester.test_results['performance_tests']:
                perf = tester.test_results['performance_tests'][0]
                print(f"  ⚡ 性能测试: 平均响应时间 {perf.get('avg_response_time', 0)}ms")
                print(f"  🚀 吞吐量: {perf.get('requests_per_second', 0)} 请求/秒")
        
        print(f"\n📝 详细报告已保存到: {report_filename}")
        print(f"📊 Excel匹配度矩阵已保存到: {results['excel_file']}")

        print("="*80)
        print("🎉 测试完成！请查看Excel文件获取详细的用户-政策匹配度矩阵")
    else:
        print("❌ 测试失败")

if __name__ == "__main__":
    main()